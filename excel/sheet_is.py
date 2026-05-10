"""Build the Income Statement worksheet."""
from __future__ import annotations
from typing import Optional

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

import config as cfg
from models.financial_data import FinancialData, IncomeStatement
from excel.styles import (
    style_title, style_col_header, style_section_header, style_subsection_header,
    style_label, style_input, style_formula, style_pct, style_subtotal, style_grand_total,
    apply_section_row, set_standard_column_widths, freeze_header, set_print_landscape,
)


def _col(idx: int) -> str:
    """1-based index to column letter (1→A, 2→B, …)."""
    return get_column_letter(idx)


def _val_or_na(v: Optional[float]) -> float | str:
    return v if v is not None else ""


def build_income_statement_sheet(wb: Workbook, data: FinancialData) -> None:
    ws = wb[cfg.SHEET_IS]
    ws.sheet_properties.tabColor = cfg.TAB_IS

    years = data.sorted_years()
    n = len(years)
    meta = data.metadata

    # ── Header rows ────────────────────────────────────────────────────────────
    # Row 1: company name title
    style_title(ws.cell(1, 1), meta.company_name)
    for c in range(2, n + 2):
        style_title(ws.cell(1, c))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n + 1)
    ws.row_dimensions[1].height = cfg.ROW_HEIGHT_TITLE

    # Row 2: subtitle
    units_label = f"Income Statement  ($ in {meta.units}, except per share)"
    style_subsection_header(ws.cell(2, 1), units_label)
    for c in range(2, n + 2):
        style_subsection_header(ws.cell(2, c))
    ws.row_dimensions[2].height = cfg.ROW_HEIGHT_HEADER

    # Row 3: year column headers
    style_col_header(ws.cell(3, 1), "")
    for i, yr in enumerate(years):
        style_col_header(ws.cell(3, i + 2), f"FY{yr}")
    ws.row_dimensions[3].height = cfg.ROW_HEIGHT_HEADER

    # Map years to column indices (col 2 = first year)
    yr_col = {yr: i + 2 for i, yr in enumerate(years)}

    row = 4

    def input_row(label: str, getter, indent: int = 1) -> int:
        nonlocal row
        style_label(ws.cell(row, 1), label, indent=indent)
        for yr in years:
            is_ = data.income_statement_for_year(yr)
            val = getter(is_) if is_ else None
            style_input(ws.cell(row, yr_col[yr]), _val_or_na(val))
        ws.row_dimensions[row].height = cfg.ROW_HEIGHT_NORMAL
        r = row; row += 1
        return r

    def formula_row(label: str, formula_fn, fmt: str = cfg.FMT_CURRENCY,
                    bold: bool = False, style_fn=None, indent: int = 1) -> int:
        nonlocal row
        style_label(ws.cell(row, 1), label, indent=indent, bold=bold)
        for yr in years:
            c = yr_col[yr]
            formula = formula_fn(c)
            if style_fn:
                style_fn(ws.cell(row, c), formula)
            else:
                style_formula(ws.cell(row, c), formula, fmt)
        ws.row_dimensions[row].height = cfg.ROW_HEIGHT_NORMAL
        r = row; row += 1
        return r

    def spacer() -> None:
        nonlocal row
        ws.row_dimensions[row].height = cfg.ROW_HEIGHT_SPACER
        row += 1

    def section(label: str) -> None:
        nonlocal row
        apply_section_row(ws, row, label, n)
        row += 1

    # ── REVENUE ────────────────────────────────────────────────────────────────
    section("Revenue")
    rev_row = input_row("Revenue", lambda is_: is_.revenue, indent=2)
    cogs_row = input_row("Cost of Revenue", lambda is_: is_.cost_of_revenue, indent=2)

    gp_row = formula_row(
        "Gross Profit",
        lambda c: f"={_col(c)}{rev_row}-{_col(c)}{cogs_row}",
        style_fn=style_subtotal,
        bold=True, indent=1,
    )
    spacer()

    # ── OPERATING EXPENSES ─────────────────────────────────────────────────────
    section("Operating Expenses")
    rd_row = input_row("Research & Development", lambda is_: is_.research_and_development, indent=2)
    sga_row = input_row("Selling, General & Administrative", lambda is_: is_.selling_general_admin, indent=2)
    other_opex_row = input_row("Other Operating Expenses", lambda is_: is_.operating_expenses_other, indent=2)

    total_opex_row = formula_row(
        "Total Operating Expenses",
        lambda c: f"=SUM({_col(c)}{rd_row}:{_col(c)}{other_opex_row})",
        style_fn=style_subtotal,
        bold=True, indent=1,
    )
    spacer()

    # ── EBITDA / EBIT ──────────────────────────────────────────────────────────
    section("EBITDA & EBIT")
    da_row = input_row("Depreciation & Amortization", lambda is_: is_.depreciation_amortization, indent=2)

    ebitda_row = formula_row(
        "EBITDA",
        lambda c: f"={_col(c)}{gp_row}-{_col(c)}{total_opex_row}+{_col(c)}{da_row}",
        style_fn=style_subtotal,
        bold=True, indent=1,
    )

    ebit_row = formula_row(
        "EBIT (Operating Income)",
        lambda c: f"={_col(c)}{ebitda_row}-{_col(c)}{da_row}",
        style_fn=style_subtotal,
        bold=True, indent=1,
    )
    spacer()

    # ── BELOW THE LINE ─────────────────────────────────────────────────────────
    section("Below the Line")
    int_exp_row = input_row("Interest Expense", lambda is_: is_.interest_expense, indent=2)
    int_inc_row = input_row("Interest Income", lambda is_: is_.interest_income, indent=2)
    other_inc_row = input_row("Other Income / (Expense)", lambda is_: is_.other_income_expense, indent=2)

    ebt_row = formula_row(
        "Pretax Income (EBT)",
        lambda c: (
            f"={_col(c)}{ebit_row}"
            f"+IF(ISNUMBER({_col(c)}{int_exp_row}),{_col(c)}{int_exp_row},0)"
            f"+IF(ISNUMBER({_col(c)}{int_inc_row}),{_col(c)}{int_inc_row},0)"
            f"+IF(ISNUMBER({_col(c)}{other_inc_row}),{_col(c)}{other_inc_row},0)"
        ),
        style_fn=style_subtotal,
        bold=True, indent=1,
    )

    tax_row = input_row("Income Tax Expense", lambda is_: is_.income_tax, indent=2)

    ni_row = formula_row(
        "Net Income",
        lambda c: f"={_col(c)}{ebt_row}-IF(ISNUMBER({_col(c)}{tax_row}),{_col(c)}{tax_row},0)",
        style_fn=style_grand_total,
        bold=True, indent=1,
    )
    spacer()

    # ── PER SHARE ──────────────────────────────────────────────────────────────
    section("Per Share Data")
    shares_basic_row = input_row("Shares Outstanding — Basic (mm)", lambda is_: is_.shares_basic, indent=2)
    shares_dil_row = input_row("Shares Outstanding — Diluted (mm)", lambda is_: is_.shares_diluted, indent=2)

    formula_row(
        "EPS — Basic",
        lambda c: f"=IF(ISNUMBER({_col(c)}{shares_basic_row}),{_col(c)}{ni_row}/{_col(c)}{shares_basic_row},\"\")",
        fmt=cfg.FMT_CURRENCY_DEC, indent=2,
    )
    formula_row(
        "EPS — Diluted",
        lambda c: f"=IF(ISNUMBER({_col(c)}{shares_dil_row}),{_col(c)}{ni_row}/{_col(c)}{shares_dil_row},\"\")",
        fmt=cfg.FMT_CURRENCY_DEC, indent=2,
    )
    spacer()

    # ── MARGIN ANALYSIS ────────────────────────────────────────────────────────
    section("Margin Analysis")
    formula_row("Gross Margin %",
                lambda c: f"=IF({_col(c)}{rev_row}<>0,{_col(c)}{gp_row}/{_col(c)}{rev_row},\"\")",
                fmt=cfg.FMT_PCT, indent=2)
    formula_row("EBITDA Margin %",
                lambda c: f"=IF({_col(c)}{rev_row}<>0,{_col(c)}{ebitda_row}/{_col(c)}{rev_row},\"\")",
                fmt=cfg.FMT_PCT, indent=2)
    formula_row("EBIT Margin %",
                lambda c: f"=IF({_col(c)}{rev_row}<>0,{_col(c)}{ebit_row}/{_col(c)}{rev_row},\"\")",
                fmt=cfg.FMT_PCT, indent=2)
    formula_row("Net Margin %",
                lambda c: f"=IF({_col(c)}{rev_row}<>0,{_col(c)}{ni_row}/{_col(c)}{rev_row},\"\")",
                fmt=cfg.FMT_PCT, indent=2)
    spacer()

    # ── YOY GROWTH ─────────────────────────────────────────────────────────────
    if n >= 2:
        section("YoY Growth")
        for label, src_row in [
            ("Revenue Growth", rev_row),
            ("Gross Profit Growth", gp_row),
            ("EBITDA Growth", ebitda_row),
            ("Net Income Growth", ni_row),
        ]:
            style_label(ws.cell(row, 1), label, indent=2)
            for i, yr in enumerate(years):
                c = yr_col[yr]
                if i == 0:
                    ws.cell(row, c).value = "n/a"
                else:
                    prev_c = yr_col[years[i - 1]]
                    style_pct(ws.cell(row, c),
                              f"=IF({_col(prev_c)}{src_row}<>0,"
                              f"({_col(c)}{src_row}-{_col(prev_c)}{src_row})/{_col(prev_c)}{src_row},\"\")")
            ws.row_dimensions[row].height = cfg.ROW_HEIGHT_NORMAL
            row += 1

    # ── Column widths & display ────────────────────────────────────────────────
    set_standard_column_widths(ws, n)
    freeze_header(ws, freeze_row=4)
    set_print_landscape(ws)

    # ── Register named ranges for cross-sheet linking ──────────────────────────
    _register_named_ranges(wb, ws, years, yr_col,
                           rev_row=rev_row, gp_row=gp_row, ebitda_row=ebitda_row,
                           ebit_row=ebit_row, da_row=da_row, ni_row=ni_row,
                           int_exp_row=int_exp_row, cogs_row=cogs_row,
                           shares_dil_row=shares_dil_row, tax_row=tax_row,
                           ebt_row=ebt_row)


def _register_named_ranges(wb: Workbook, ws: Worksheet, years: list[str],
                            yr_col: dict, **row_map) -> None:
    from openpyxl.workbook.defined_name import DefinedName
    sn = f"'{cfg.SHEET_IS}'"

    for i, yr in enumerate(years):
        c = get_column_letter(yr_col[yr])
        suffix = f"_FY{i + 1}"

        for name, row_key in [
            ("IS_Revenue", "rev_row"),
            ("IS_GrossProfit", "gp_row"),
            ("IS_EBITDA", "ebitda_row"),
            ("IS_EBIT", "ebit_row"),
            ("IS_DA", "da_row"),
            ("IS_NetIncome", "ni_row"),
            ("IS_IntExp", "int_exp_row"),
            ("IS_COGS", "cogs_row"),
            ("IS_SharesDil", "shares_dil_row"),
            ("IS_Tax", "tax_row"),
            ("IS_EBT", "ebt_row"),
        ]:
            r = row_map[row_key]
            ref = f"{sn}!${c}${r}"
            try:
                wb.defined_names[name + suffix] = DefinedName(name + suffix, attr_text=ref)
            except Exception:
                pass  # named range already exists (shouldn't happen on new workbook)
