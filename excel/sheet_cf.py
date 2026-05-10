"""Build the Cash Flow Statement worksheet with cross-sheet links to IS and BS."""
from __future__ import annotations
from typing import Optional

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.styles.fills import FILL_SOLID
from openpyxl.formatting.rule import FormulaRule

import config as cfg
from models.financial_data import FinancialData
from excel.styles import (
    style_title, style_col_header, style_subsection_header,
    style_label, style_input, style_link, style_subtotal, style_grand_total,
    apply_section_row, apply_subsection_row, set_standard_column_widths,
    freeze_header, set_print_landscape,
)


def _col(idx: int) -> str:
    return get_column_letter(idx)


def _val(v: Optional[float]) -> float | str:
    return v if v is not None else ""


def build_cash_flow_sheet(wb: Workbook, data: FinancialData) -> None:
    ws = wb[cfg.SHEET_CF]
    ws.sheet_properties.tabColor = cfg.TAB_CF

    years = data.sorted_years()
    n = len(years)
    meta = data.metadata

    # ── Header ─────────────────────────────────────────────────────────────────
    style_title(ws.cell(1, 1), meta.company_name)
    for c in range(2, n + 2):
        style_title(ws.cell(1, c))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n + 1)
    ws.row_dimensions[1].height = cfg.ROW_HEIGHT_TITLE

    style_subsection_header(ws.cell(2, 1), f"Cash Flow Statement  ($ in {meta.units})")
    for c in range(2, n + 2):
        style_subsection_header(ws.cell(2, c))
    ws.row_dimensions[2].height = cfg.ROW_HEIGHT_HEADER

    style_col_header(ws.cell(3, 1), "")
    for i, yr in enumerate(years):
        style_col_header(ws.cell(3, i + 2), f"FY{yr}")
    ws.row_dimensions[3].height = cfg.ROW_HEIGHT_HEADER

    yr_col = {yr: i + 2 for i, yr in enumerate(years)}
    row = 4

    def inp(label: str, getter, indent: int = 2) -> int:
        nonlocal row
        style_label(ws.cell(row, 1), label, indent=indent)
        for yr in years:
            cf = data.cash_flow_for_year(yr)
            val = getter(cf) if cf else None
            style_input(ws.cell(row, yr_col[yr]), _val(val))
        ws.row_dimensions[row].height = cfg.ROW_HEIGHT_NORMAL
        r = row; row += 1
        return r

    def link_row(label: str, named_range_base: str, fmt: str = cfg.FMT_CURRENCY, indent: int = 2) -> int:
        """Row that links to a named range on IS or BS."""
        nonlocal row
        style_label(ws.cell(row, 1), label, indent=indent)
        for i, yr in enumerate(years):
            c = yr_col[yr]
            nr = f"{named_range_base}_FY{i + 1}"
            style_link(ws.cell(row, c), f"={nr}", fmt)
        ws.row_dimensions[row].height = cfg.ROW_HEIGHT_NORMAL
        r = row; row += 1
        return r

    def subtotal_sum(label: str, first: int, last: int, indent: int = 1) -> int:
        nonlocal row
        style_label(ws.cell(row, 1), label, indent=indent, bold=True)
        for yr in years:
            c = yr_col[yr]
            style_subtotal(ws.cell(row, c),
                           f"=SUM({_col(c)}{first}:{_col(c)}{last})")
        ws.row_dimensions[row].height = cfg.ROW_HEIGHT_NORMAL
        r = row; row += 1
        return r

    def spacer() -> None:
        nonlocal row
        ws.row_dimensions[row].height = cfg.ROW_HEIGHT_SPACER
        row += 1

    def section(label: str) -> None:
        nonlocal row
        apply_section_row(ws, row, label, n)
        row += 1

    # ── OPERATING ACTIVITIES ───────────────────────────────────────────────────
    section("Operating Activities")
    ni_row = link_row("Net Income", "IS_NetIncome")
    da_row = link_row("Depreciation & Amortization", "IS_DA")
    sbc_row = inp("Stock-based Compensation", lambda cf: cf.stock_based_compensation)
    wc_row = inp("Changes in Working Capital", lambda cf: cf.changes_in_working_capital)
    ar_chg_row = inp("  Change in Accounts Receivable", lambda cf: cf.change_accounts_receivable, indent=3)
    inv_chg_row = inp("  Change in Inventory", lambda cf: cf.change_inventory, indent=3)
    ap_chg_row = inp("  Change in Accounts Payable", lambda cf: cf.change_accounts_payable, indent=3)
    other_wc_row = inp("  Change in Other Working Capital", lambda cf: cf.change_other_working_capital, indent=3)
    other_op_row = inp("Other Operating Activities", lambda cf: cf.other_operating_activities)
    cfo_row = subtotal_sum("Net Cash from Operations", ni_row, other_op_row)
    spacer()

    # ── INVESTING ACTIVITIES ───────────────────────────────────────────────────
    section("Investing Activities")
    capex_row = inp("Capital Expenditures", lambda cf: cf.capital_expenditures)
    acq_row = inp("Acquisitions, net", lambda cf: cf.acquisitions)
    buy_inv_row = inp("Purchases of Investments", lambda cf: cf.purchases_investments)
    sell_inv_row = inp("Proceeds from Sales of Investments", lambda cf: cf.sales_investments)
    other_inv_row = inp("Other Investing Activities", lambda cf: cf.other_investing)
    cfi_row = subtotal_sum("Net Cash from Investing", capex_row, other_inv_row)
    spacer()

    # ── FINANCING ACTIVITIES ───────────────────────────────────────────────────
    section("Financing Activities")
    debt_iss_row = inp("Proceeds from Debt Issuance", lambda cf: cf.debt_issuance)
    debt_rep_row = inp("Debt Repayments", lambda cf: cf.debt_repayment)
    divs_row = inp("Dividends Paid", lambda cf: cf.dividends_paid)
    buyback_row = inp("Share Repurchases", lambda cf: cf.share_repurchases)
    share_iss_row = inp("Share Issuances / ESOP", lambda cf: cf.share_issuance)
    other_fin_row = inp("Other Financing Activities", lambda cf: cf.other_financing)
    cff_row = subtotal_sum("Net Cash from Financing", debt_iss_row, other_fin_row)
    spacer()

    # ── RECONCILIATION ─────────────────────────────────────────────────────────
    section("Cash Reconciliation")

    style_label(ws.cell(row, 1), "Net Change in Cash", indent=2, bold=True)
    for yr in years:
        c = yr_col[yr]
        style_subtotal(ws.cell(row, c),
                       f"={_col(c)}{cfo_row}+{_col(c)}{cfi_row}+{_col(c)}{cff_row}")
    ws.row_dimensions[row].height = cfg.ROW_HEIGHT_NORMAL
    net_change_row = row; row += 1

    beg_cash_row = inp("Beginning Cash", lambda cf: cf.beginning_cash)
    end_cash_cf_row = inp("Ending Cash (per CF)", lambda cf: cf.ending_cash)

    style_label(ws.cell(row, 1), "Ending Cash (per Balance Sheet)", indent=2)
    for i, yr in enumerate(years):
        c = yr_col[yr]
        nr = f"BS_Cash_FY{i + 1}"
        style_link(ws.cell(row, c), f"={nr}")
    ws.row_dimensions[row].height = cfg.ROW_HEIGHT_NORMAL
    end_cash_bs_row = row; row += 1
    spacer()

    # ── CF / BS Cash Check ─────────────────────────────────────────────────────
    apply_subsection_row(ws, row, "Cash Reconciliation Check", n)
    row += 1
    style_label(ws.cell(row, 1), "CF Ending Cash = BS Cash?", indent=2, bold=True)
    for yr in years:
        c = yr_col[yr]
        cell = ws.cell(row, c)
        cell.value = (
            f'=IF(OR({_col(c)}{end_cash_cf_row}="",{_col(c)}{end_cash_bs_row}=""),"n/a",'
            f'IF(ABS({_col(c)}{end_cash_cf_row}-{_col(c)}{end_cash_bs_row})<1,"✓ OK","MISMATCH"))'
        )
        ok_fill = PatternFill(fill_type=FILL_SOLID, fgColor=cfg.FILL_CHECK_OK)
        fail_fill = PatternFill(fill_type=FILL_SOLID, fgColor=cfg.FILL_CHECK_FAIL)
        ref = f"{_col(c)}{row}"
        ws.conditional_formatting.add(ref, FormulaRule(formula=[f'LEFT({ref},1)="✓"'], fill=ok_fill))
        ws.conditional_formatting.add(ref, FormulaRule(formula=[f'LEFT({ref},1)="M"'], fill=fail_fill))
    ws.row_dimensions[row].height = cfg.ROW_HEIGHT_NORMAL
    row += 1

    # ── Column widths & display ────────────────────────────────────────────────
    set_standard_column_widths(ws, n)
    freeze_header(ws, freeze_row=4)
    set_print_landscape(ws)

    # ── Named ranges ──────────────────────────────────────────────────────────
    _register_named_ranges(wb, ws, years, yr_col,
                           cfo_row=cfo_row, cfi_row=cfi_row, cff_row=cff_row,
                           capex_row=capex_row)


def _register_named_ranges(wb: Workbook, ws, years, yr_col, **row_map) -> None:
    from openpyxl.workbook.defined_name import DefinedName
    sn = f"'{cfg.SHEET_CF}'"
    for i, yr in enumerate(years):
        c = get_column_letter(yr_col[yr])
        suffix = f"_FY{i + 1}"
        for name, row_key in [
            ("CF_CFO", "cfo_row"),
            ("CF_CFI", "cfi_row"),
            ("CF_CFF", "cff_row"),
            ("CF_CapEx", "capex_row"),
        ]:
            r = row_map[row_key]
            ref = f"{sn}!${c}${r}"
            try:
                wb.defined_names[name + suffix] = DefinedName(name + suffix, attr_text=ref)
            except Exception:
                pass
