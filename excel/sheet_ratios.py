"""Build the Ratios, KPIs, and Working Capital Schedule worksheet."""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

import config as cfg
from models.financial_data import FinancialData
from excel.styles import (
    style_title, style_col_header, style_subsection_header,
    style_label, style_formula, style_pct, style_link, style_subtotal,
    apply_section_row, set_standard_column_widths, freeze_header, set_print_landscape,
)


def _col(idx: int) -> str:
    return get_column_letter(idx)


def build_ratios_sheet(wb: Workbook, data: FinancialData) -> None:
    ws = wb[cfg.SHEET_RATIOS]
    ws.sheet_properties.tabColor = cfg.TAB_RATIOS

    years = data.sorted_years()
    n = len(years)
    meta = data.metadata

    # ── Header ─────────────────────────────────────────────────────────────────
    style_title(ws.cell(1, 1), meta.company_name)
    for c in range(2, n + 2):
        style_title(ws.cell(1, c))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n + 1)
    ws.row_dimensions[1].height = cfg.ROW_HEIGHT_TITLE

    style_subsection_header(ws.cell(2, 1), "Ratios & KPIs")
    for c in range(2, n + 2):
        style_subsection_header(ws.cell(2, c))
    ws.row_dimensions[2].height = cfg.ROW_HEIGHT_HEADER

    style_col_header(ws.cell(3, 1), "")
    for i, yr in enumerate(years):
        style_col_header(ws.cell(3, i + 2), f"FY{yr}")
    ws.row_dimensions[3].height = cfg.ROW_HEIGHT_HEADER

    yr_col = {yr: i + 2 for i, yr in enumerate(years)}
    row = 4

    # Helper: write a ratio row using named range references
    def ratio_pct(label: str, formula_fn, indent: int = 2) -> int:
        nonlocal row
        style_label(ws.cell(row, 1), label, indent=indent)
        for i, yr in enumerate(years):
            c = yr_col[yr]
            style_pct(ws.cell(row, c), formula_fn(i + 1, c))
        ws.row_dimensions[row].height = cfg.ROW_HEIGHT_NORMAL
        r = row; row += 1
        return r

    def ratio_num(label: str, formula_fn, fmt: str = cfg.FMT_CURRENCY, indent: int = 2) -> int:
        nonlocal row
        style_label(ws.cell(row, 1), label, indent=indent)
        for i, yr in enumerate(years):
            c = yr_col[yr]
            style_formula(ws.cell(row, c), formula_fn(i + 1, c), fmt)
        ws.row_dimensions[row].height = cfg.ROW_HEIGHT_NORMAL
        r = row; row += 1
        return r

    def ratio_multiple(label: str, formula_fn, indent: int = 2) -> int:
        return ratio_num(label, formula_fn, fmt=cfg.FMT_MULTIPLE, indent=indent)

    def spacer() -> None:
        nonlocal row
        ws.row_dimensions[row].height = cfg.ROW_HEIGHT_SPACER
        row += 1

    def section(label: str) -> None:
        nonlocal row
        apply_section_row(ws, row, label, n)
        row += 1

    # ── Helper: safe division formula ─────────────────────────────────────────
    def div(num: str, denom: str) -> str:
        return f"=IF(AND(ISNUMBER({denom}),{denom}<>0),{num}/{denom},\"\")"

    # ── A. PROFITABILITY ───────────────────────────────────────────────────────
    section("A. Profitability Ratios")

    ratio_pct("Gross Margin %",
              lambda fy, c: div(f"IS_GrossProfit_FY{fy}", f"IS_Revenue_FY{fy}"))
    ratio_pct("EBITDA Margin %",
              lambda fy, c: div(f"IS_EBITDA_FY{fy}", f"IS_Revenue_FY{fy}"))
    ratio_pct("EBIT Margin %",
              lambda fy, c: div(f"IS_EBIT_FY{fy}", f"IS_Revenue_FY{fy}"))
    ratio_pct("Net Profit Margin %",
              lambda fy, c: div(f"IS_NetIncome_FY{fy}", f"IS_Revenue_FY{fy}"))
    ratio_pct("Return on Assets (ROA)",
              lambda fy, c: div(f"IS_NetIncome_FY{fy}", f"BS_TotalAssets_FY{fy}"))
    ratio_pct("Return on Equity (ROE)",
              lambda fy, c: div(f"IS_NetIncome_FY{fy}", f"BS_TotalEquity_FY{fy}"))
    ratio_pct("Return on Invested Capital (ROIC)",
              lambda fy, c: (
                  f"=IF(AND(ISNUMBER(IS_EBIT_FY{fy}),ISNUMBER(IS_Tax_FY{fy}),ISNUMBER(IS_EBT_FY{fy})),"
                  f"IS_EBIT_FY{fy}*(1-IF(IS_EBT_FY{fy}<>0,IS_Tax_FY{fy}/IS_EBT_FY{fy},0))/"
                  f"(BS_TotalEquity_FY{fy}+IF(ISNUMBER(BS_LTD_FY{fy}),BS_LTD_FY{fy},0)"
                  f"+IF(ISNUMBER(BS_STD_FY{fy}),BS_STD_FY{fy},0)-BS_Cash_FY{fy}),\"\")"
              ))
    spacer()

    # ── B. LIQUIDITY ───────────────────────────────────────────────────────────
    section("B. Liquidity Ratios")

    ratio_multiple("Current Ratio",
                   lambda fy, c: div(f"BS_TCA_FY{fy}", f"BS_TCL_FY{fy}"))
    ratio_multiple("Quick Ratio",
                   lambda fy, c: (
                       f"=IF(AND(ISNUMBER(BS_TCL_FY{fy}),BS_TCL_FY{fy}<>0),"
                       f"(BS_Cash_FY{fy}+IF(ISNUMBER(BS_STI_FY{fy}),BS_STI_FY{fy},0)"
                       f"+IF(ISNUMBER(BS_AR_FY{fy}),BS_AR_FY{fy},0))/BS_TCL_FY{fy},\"\")"
                   ))
    ratio_multiple("Cash Ratio",
                   lambda fy, c: div(f"BS_Cash_FY{fy}", f"BS_TCL_FY{fy}"))
    ratio_num("Net Working Capital ($mm)",
              lambda fy, c: (
                  f"=IF(AND(ISNUMBER(BS_TCA_FY{fy}),ISNUMBER(BS_TCL_FY{fy})),"
                  f"BS_TCA_FY{fy}-BS_TCL_FY{fy},\"\")"
              ))
    spacer()

    # ── C. LEVERAGE ────────────────────────────────────────────────────────────
    section("C. Leverage Ratios")

    ratio_multiple("Total Debt / Equity",
                   lambda fy, c: (
                       f"=IF(AND(ISNUMBER(BS_TotalEquity_FY{fy}),BS_TotalEquity_FY{fy}<>0),"
                       f"(IF(ISNUMBER(BS_LTD_FY{fy}),BS_LTD_FY{fy},0)"
                       f"+IF(ISNUMBER(BS_STD_FY{fy}),BS_STD_FY{fy},0))/BS_TotalEquity_FY{fy},\"\")"
                   ))
    ratio_num("Net Debt ($mm)",
              lambda fy, c: (
                  f"=IF(ISNUMBER(BS_Cash_FY{fy}),"
                  f"IF(ISNUMBER(BS_LTD_FY{fy}),BS_LTD_FY{fy},0)"
                  f"+IF(ISNUMBER(BS_STD_FY{fy}),BS_STD_FY{fy},0)-BS_Cash_FY{fy},\"\")"
              ))
    ratio_multiple("Net Debt / EBITDA",
                   lambda fy, c: (
                       f"=IF(AND(ISNUMBER(IS_EBITDA_FY{fy}),IS_EBITDA_FY{fy}<>0,"
                       f"ISNUMBER(BS_Cash_FY{fy})),"
                       f"(IF(ISNUMBER(BS_LTD_FY{fy}),BS_LTD_FY{fy},0)"
                       f"+IF(ISNUMBER(BS_STD_FY{fy}),BS_STD_FY{fy},0)-BS_Cash_FY{fy})"
                       f"/IS_EBITDA_FY{fy},\"\")"
                   ))
    ratio_multiple("Interest Coverage (EBIT / Int. Exp.)",
                   lambda fy, c: (
                       f"=IF(AND(ISNUMBER(IS_IntExp_FY{fy}),IS_IntExp_FY{fy}<>0,ISNUMBER(IS_EBIT_FY{fy})),"
                       f"IS_EBIT_FY{fy}/ABS(IS_IntExp_FY{fy}),\"\")"
                   ))
    ratio_pct("Total Debt / Total Assets",
              lambda fy, c: (
                  f"=IF(AND(ISNUMBER(BS_TotalAssets_FY{fy}),BS_TotalAssets_FY{fy}<>0),"
                  f"(IF(ISNUMBER(BS_LTD_FY{fy}),BS_LTD_FY{fy},0)"
                  f"+IF(ISNUMBER(BS_STD_FY{fy}),BS_STD_FY{fy},0))/BS_TotalAssets_FY{fy},\"\")"
              ))
    spacer()

    # ── D. WORKING CAPITAL SCHEDULE ────────────────────────────────────────────
    section("D. Working Capital Schedule")

    ratio_num("Days Sales Outstanding (DSO)",
              lambda fy, c: (
                  f"=IF(AND(ISNUMBER(BS_AR_FY{fy}),ISNUMBER(IS_Revenue_FY{fy}),IS_Revenue_FY{fy}<>0),"
                  f"BS_AR_FY{fy}/IS_Revenue_FY{fy}*365,\"\")"
              ), fmt=cfg.FMT_CURRENCY_DEC)
    ratio_num("Days Inventory Outstanding (DIO)",
              lambda fy, c: (
                  f"=IF(AND(ISNUMBER(BS_Inventory_FY{fy}),ISNUMBER(IS_COGS_FY{fy}),IS_COGS_FY{fy}<>0),"
                  f"BS_Inventory_FY{fy}/IS_COGS_FY{fy}*365,\"\")"
              ), fmt=cfg.FMT_CURRENCY_DEC)
    ratio_num("Days Payable Outstanding (DPO)",
              lambda fy, c: (
                  f"=IF(AND(ISNUMBER(BS_AP_FY{fy}),ISNUMBER(IS_COGS_FY{fy}),IS_COGS_FY{fy}<>0),"
                  f"BS_AP_FY{fy}/IS_COGS_FY{fy}*365,\"\")"
              ), fmt=cfg.FMT_CURRENCY_DEC)
    ratio_num("Cash Conversion Cycle (DSO + DIO - DPO)",
              lambda fy, c: (
                  f"=IF(AND(ISNUMBER(BS_AR_FY{fy}),ISNUMBER(IS_Revenue_FY{fy}),IS_Revenue_FY{fy}<>0),"
                  f"IF(ISNUMBER(BS_AR_FY{fy}),BS_AR_FY{fy}/IS_Revenue_FY{fy}*365,0)"
                  f"+IF(AND(ISNUMBER(BS_Inventory_FY{fy}),ISNUMBER(IS_COGS_FY{fy}),IS_COGS_FY{fy}<>0),"
                  f"BS_Inventory_FY{fy}/IS_COGS_FY{fy}*365,0)"
                  f"-IF(AND(ISNUMBER(BS_AP_FY{fy}),ISNUMBER(IS_COGS_FY{fy}),IS_COGS_FY{fy}<>0),"
                  f"BS_AP_FY{fy}/IS_COGS_FY{fy}*365,0),\"\")"
              ), fmt=cfg.FMT_CURRENCY_DEC)

    nwc_row = ratio_num("Net Working Capital ($mm)",
                        lambda fy, c: (
                            f"=IF(AND(ISNUMBER(BS_TCA_FY{fy}),ISNUMBER(BS_TCL_FY{fy})),"
                            f"BS_TCA_FY{fy}-BS_TCL_FY{fy},\"\")"
                        ))
    ratio_pct("NWC as % of Revenue",
              lambda fy, c: (
                  f"=IF(AND(ISNUMBER(BS_TCA_FY{fy}),ISNUMBER(IS_Revenue_FY{fy}),IS_Revenue_FY{fy}<>0),"
                  f"(BS_TCA_FY{fy}-BS_TCL_FY{fy})/IS_Revenue_FY{fy},\"\")"
              ))
    if n >= 2:
        style_label(ws.cell(row, 1), "YoY Change in NWC ($mm)", indent=2)
        for i, yr in enumerate(years):
            c = yr_col[yr]
            if i == 0:
                ws.cell(row, c).value = "n/a"
            else:
                prev_fy = i  # previous FY index (1-based)
                cur_fy = i + 1
                style_formula(ws.cell(row, c),
                              f"=IF(AND(ISNUMBER(BS_TCA_FY{cur_fy}),ISNUMBER(BS_TCA_FY{prev_fy})),"
                              f"(BS_TCA_FY{cur_fy}-BS_TCL_FY{cur_fy})-(BS_TCA_FY{prev_fy}-BS_TCL_FY{prev_fy}),\"\")")
        ws.row_dimensions[row].height = cfg.ROW_HEIGHT_NORMAL
        row += 1
    spacer()

    # ── E. PER SHARE ───────────────────────────────────────────────────────────
    section("E. Per Share Data")

    ratio_num("EPS (Diluted)",
              lambda fy, c: div(f"IS_NetIncome_FY{fy}", f"IS_SharesDil_FY{fy}"),
              fmt=cfg.FMT_CURRENCY_DEC)
    ratio_num("Book Value per Share",
              lambda fy, c: div(f"BS_TotalEquity_FY{fy}", f"IS_SharesDil_FY{fy}"),
              fmt=cfg.FMT_CURRENCY_DEC)
    ratio_num("Revenue per Share",
              lambda fy, c: div(f"IS_Revenue_FY{fy}", f"IS_SharesDil_FY{fy}"),
              fmt=cfg.FMT_CURRENCY_DEC)

    # ── Column widths & display ────────────────────────────────────────────────
    set_standard_column_widths(ws, n)
    freeze_header(ws, freeze_row=4)
    set_print_landscape(ws)
