from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.styles.fills import FILL_SOLID
from openpyxl.worksheet.worksheet import Worksheet
import config as cfg


def _side(style: str = "thin") -> Side:
    return Side(style=style)


def thin_border() -> Border:
    return Border(
        left=_side(), right=_side(), top=_side(), bottom=_side()
    )


def bottom_border(style: str = "medium") -> Border:
    return Border(bottom=_side(style))


def thick_top_border() -> Border:
    return Border(top=_side("medium"))


def double_bottom_border() -> Border:
    return Border(
        top=_side("thin"),
        bottom=_side("double"),
    )


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type=FILL_SOLID, fgColor=hex_color)


def _font(
    bold: bool = False,
    italic: bool = False,
    size: int = cfg.FONT_SIZE_NORMAL,
    color: str = cfg.COLOR_FORMULA,
    name: str = cfg.FONT_NAME,
) -> Font:
    return Font(bold=bold, italic=italic, size=size, color=color, name=name)


def center_align(wrap: bool = False) -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def left_align(indent: int = 0, wrap: bool = False) -> Alignment:
    return Alignment(horizontal="left", vertical="center", indent=indent, wrap_text=wrap)


def right_align() -> Alignment:
    return Alignment(horizontal="right", vertical="center")


# ── Style applicators ─────────────────────────────────────────────────────────

def style_title(cell, text: str | None = None) -> None:
    if text is not None:
        cell.value = text
    cell.font = _font(bold=True, size=cfg.FONT_SIZE_TITLE, color=cfg.COLOR_HEADER_TEXT)
    cell.fill = _fill(cfg.FILL_SECTION_HEADER)
    cell.alignment = left_align(indent=1)


def style_section_header(cell, text: str | None = None) -> None:
    if text is not None:
        cell.value = text
    cell.font = _font(bold=True, size=cfg.FONT_SIZE_NORMAL, color=cfg.COLOR_HEADER_TEXT)
    cell.fill = _fill(cfg.FILL_SECTION_HEADER)
    cell.alignment = left_align(indent=1)


def style_subsection_header(cell, text: str | None = None) -> None:
    if text is not None:
        cell.value = text
    cell.font = _font(bold=True, color=cfg.COLOR_FORMULA)
    cell.fill = _fill(cfg.FILL_SUBSECTION)
    cell.alignment = left_align(indent=1)


def style_col_header(cell, text: str | None = None) -> None:
    if text is not None:
        cell.value = text
    cell.font = _font(bold=True, color=cfg.COLOR_HEADER_TEXT)
    cell.fill = _fill(cfg.FILL_SECTION_HEADER)
    cell.alignment = center_align()


def style_input(cell, value=None, fmt: str = cfg.FMT_CURRENCY) -> None:
    if value is not None:
        cell.value = value
    cell.font = _font(color=cfg.COLOR_INPUT)
    cell.number_format = fmt
    cell.alignment = right_align()


def style_formula(cell, formula: str | None = None, fmt: str = cfg.FMT_CURRENCY) -> None:
    if formula is not None:
        cell.value = formula
    cell.font = _font(color=cfg.COLOR_FORMULA)
    cell.number_format = fmt
    cell.alignment = right_align()


def style_link(cell, formula: str | None = None, fmt: str = cfg.FMT_CURRENCY) -> None:
    """Cross-sheet link cell — green font."""
    if formula is not None:
        cell.value = formula
    cell.font = _font(color=cfg.COLOR_LINK)
    cell.number_format = fmt
    cell.alignment = right_align()


def style_subtotal(cell, formula: str | None = None, fmt: str = cfg.FMT_CURRENCY) -> None:
    if formula is not None:
        cell.value = formula
    cell.font = _font(bold=True, color=cfg.COLOR_FORMULA)
    cell.fill = _fill(cfg.FILL_TOTAL)
    cell.number_format = fmt
    cell.alignment = right_align()
    cell.border = thick_top_border()


def style_grand_total(cell, formula: str | None = None, fmt: str = cfg.FMT_CURRENCY) -> None:
    if formula is not None:
        cell.value = formula
    cell.font = _font(bold=True, color=cfg.COLOR_FORMULA)
    cell.fill = _fill(cfg.FILL_GRAND_TOTAL)
    cell.number_format = fmt
    cell.alignment = right_align()
    cell.border = double_bottom_border()


def style_label(cell, text: str | None = None, indent: int = 0, bold: bool = False) -> None:
    if text is not None:
        cell.value = text
    cell.font = _font(bold=bold)
    cell.alignment = left_align(indent=indent)


def style_pct(cell, formula: str | None = None) -> None:
    if formula is not None:
        cell.value = formula
    cell.font = _font(color=cfg.COLOR_FORMULA)
    cell.number_format = cfg.FMT_PCT
    cell.alignment = right_align()


def style_pct_subtotal(cell, formula: str | None = None) -> None:
    if formula is not None:
        cell.value = formula
    cell.font = _font(bold=True)
    cell.fill = _fill(cfg.FILL_TOTAL)
    cell.number_format = cfg.FMT_PCT
    cell.alignment = right_align()
    cell.border = thick_top_border()


def apply_section_row(ws: Worksheet, row: int, label: str, n_cols: int) -> None:
    """Apply section header style across label + data columns."""
    cell = ws.cell(row=row, column=1)
    style_section_header(cell, label)
    for c in range(2, n_cols + 2):
        style_section_header(ws.cell(row=row, column=c))
    ws.row_dimensions[row].height = cfg.ROW_HEIGHT_HEADER


def apply_subsection_row(ws: Worksheet, row: int, label: str, n_cols: int) -> None:
    cell = ws.cell(row=row, column=1)
    style_subsection_header(cell, label)
    for c in range(2, n_cols + 2):
        style_subsection_header(ws.cell(row=row, column=c))
    ws.row_dimensions[row].height = cfg.ROW_HEIGHT_HEADER


def set_standard_column_widths(ws: Worksheet, n_year_cols: int) -> None:
    ws.column_dimensions["A"].width = cfg.COL_WIDTH_LABEL
    cols = "BCDEFGHIJ"
    for i in range(n_year_cols):
        ws.column_dimensions[cols[i]].width = cfg.COL_WIDTH_YEAR


def freeze_header(ws: Worksheet, freeze_row: int = 4) -> None:
    ws.freeze_panes = ws.cell(row=freeze_row, column=2)


def set_print_landscape(ws: Worksheet) -> None:
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.gridLines = False
