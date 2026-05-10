"""Build the Balance Sheet worksheet."""
from __future__ import annotations
from typing import Optional

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.styles.fills import FILL_SOLID
from openpyxl.formatting.rule import FormulaRule

import config as cfg
from models.financial_data import FinancialData
from excel.styles import (
    style_title, style_col_header, style_section_header, style_subsection_header,
    style_label, style_input, style_formula, style_subtotal, style_grand_total,
    apply_section_row, apply_subsection_row, set_standard_column_widths,
    freeze_header, set_print_landscape,
)


def _col(idx: int) -> str:
    return get_column_letter(idx)


def _val(v: Optional[float]) -> float | str:
    return v if v is not None else ""


def build_balance_sheet_sheet(wb: Workbook, data: FinancialData) -> None:
    ws = wb[cfg.SHEET_BS]
    ws.sheet_properties.tabColor = cfg.TAB_BS

    years = data.sorted_years()
    n = len(years)
    meta = data.metadata

    # ── Header ─────────────────────────────────────────────────────────────────
    style_title(ws.cell(1, 1), meta.company_name)
    for c in range(2, n + 2):
        style_title(ws.cell(1, c))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n + 1)
    ws.row_dimensions[1].height = cfg.ROW_HEIGHT_TITLE

    style_subsection_header(ws.cell(2, 1), f"Balance Sheet  ($ in {meta.units})")
    for c in range(2, n + 2):
        style_subsection_header(ws.cell(2, c))
    ws.row_dimensions[2].height = cfg.ROW_HEIGHT_HEADER

    style_col_header(ws.cell(3, 1), "")
    for i, yr in enumerate(years):
        style_col_header(ws.cell(3, i + 2), f"FY{yr}")
    ws.row_dimensions[3].height = cfg.ROW_HEIGHT_HEADER

    yr_col = {yr: i + 2 for i, yr in enumerate(years)}
    row = 4

    def inp(label: str, getter, indent: int = 2) -> int:
        nonlocal row
        style_label(ws.cell(row, 1), label, indent=indent)
        for yr in years:
            bs = data.balance_sheet_for_year(yr)
            val = getter(bs) if bs else None
            style_input(ws.cell(row, yr_col[yr]), _val(val))
        ws.row_dimensions[row].height = cfg.ROW_HEIGHT_NORMAL
        r = row; row += 1
        return r

    def subtotal_sum(label: str, first: int, last: int, indent: int = 1) -> int:
        nonlocal row
        style_label(ws.cell(row, 1), label, indent=indent, bold=True)
        for yr in years:
            c = yr_col[yr]
            style_subtotal(ws.cell(row, c),
                           f"=SUM({_col(c)}{first}:{_col(c)}{last})")
        ws.row_dimensions[row].height = cfg.ROW_HEIGHT_NORMAL
        r = row; row += 1
        return r

    def grand_sum(label: str, addends: list[int], indent: int = 1) -> int:
        nonlocal row
        style_label(ws.cell(row, 1), label, indent=indent, bold=True)
        for yr in years:
            c = yr_col[yr]
            parts = "+".join(f"{_col(c)}{r}" for r in addends)
            style_grand_total(ws.cell(row, c), f"={parts}")
        ws.row_dimensions[row].height = cfg.ROW_HEIGHT_NORMAL
        r = row; row += 1
        return r

    def spacer() -> None:
        nonlocal row
        ws.row_dimensions[row].height = cfg.ROW_HEIGHT_SPACER
        row += 1

    def section(label: str) -> None:
        nonlocal row
        apply_section_row(ws, row, label, n)
        row += 1

    def subsection(label: str) -> None:
        nonlocal row
        apply_subsection_row(ws, row, label, n)
        row += 1

    # ── ASSETS ─────────────────────────────────────────────────────────────────
    section("ASSETS")
    subsection("Current Assets")
    cash_row = inp("Cash & Cash Equivalents", lambda bs: bs.cash_and_equivalents)
    sti_row = inp("Short-term Investments", lambda bs: bs.short_term_investments)
    ar_row = inp("Accounts Receivable, net", lambda bs: bs.accounts_receivable)
    inv_row = inp("Inventory", lambda bs: bs.inventory)
    prepaid_row = inp("Prepaid & Other Current Assets", lambda bs: bs.prepaid_other_current)
    tca_row = subtotal_sum("Total Current Assets", cash_row, prepaid_row)
    spacer()

    subsection("Non-current Assets")
    ppe_gross_row = inp("PP&E, Gross", lambda bs: bs.ppe_gross)
    accum_dep_row = inp("Accumulated Depreciation", lambda bs: bs.accumulated_depreciation)
    ppe_net_row = inp("PP&E, Net", lambda bs: bs.ppe_net)
    goodwill_row = inp("Goodwill", lambda bs: bs.goodwill)
    intang_row = inp("Intangible Assets, net", lambda bs: bs.intangibles)
    other_nca_row = inp("Other Non-current Assets", lambda bs: bs.other_noncurrent_assets)
    tnca_row = subtotal_sum("Total Non-current Assets", ppe_net_row, other_nca_row)
    spacer()

    total_assets_row = grand_sum("Total Assets", [tca_row, tnca_row])
    spacer()

    # ── LIABILITIES ────────────────────────────────────────────────────────────
    section("LIABILITIES")
    subsection("Current Liabilities")
    ap_row = inp("Accounts Payable", lambda bs: bs.accounts_payable)
    std_row = inp("Short-term Debt / Current Portion LTD", lambda bs: bs.short_term_debt)
    accrued_row = inp("Accrued Liabilities", lambda bs: bs.accrued_liabilities)
    def_rev_row = inp("Deferred Revenue (Current)", lambda bs: bs.deferred_revenue_current)
    other_cl_row = inp("Other Current Liabilities", lambda bs: bs.other_current_liabilities)
    tcl_row = subtotal_sum("Total Current Liabilities", ap_row, other_cl_row)
    spacer()

    subsection("Non-current Liabilities")
    ltd_row = inp("Long-term Debt", lambda bs: bs.long_term_debt)
    def_tax_row = inp("Deferred Tax Liabilities", lambda bs: bs.deferred_tax_liabilities)
    other_ncl_row = inp("Other Non-current Liabilities", lambda bs: bs.other_noncurrent_liabilities)
    tncl_row = subtotal_sum("Total Non-current Liabilities", ltd_row, other_ncl_row)
    spacer()

    total_liab_row = grand_sum("Total Liabilities", [tcl_row, tncl_row])
    spacer()

    # ── EQUITY ─────────────────────────────────────────────────────────────────
    section("SHAREHOLDERS' EQUITY")
    cs_row = inp("Common Stock & APIC", lambda bs: (bs.common_stock or 0) + (bs.additional_paid_in_capital or 0))
    re_row = inp("Retained Earnings", lambda bs: bs.retained_earnings)
    aoci_row = inp("Accumulated Other Comprehensive Income/(Loss)", lambda bs: bs.accumulated_other_comprehensive)
    treasury_row = inp("Treasury Stock", lambda bs: bs.treasury_stock)
    total_equity_row = subtotal_sum("Total Shareholders' Equity", cs_row, treasury_row)
    spacer()

    total_le_row = grand_sum("Total Liabilities & Equity", [total_liab_row, total_equity_row])
    spacer()

    # ── BALANCE CHECK ──────────────────────────────────────────────────────────
    apply_subsection_row(ws, row, "Balance Check", n)
    row += 1
    style_label(ws.cell(row, 1), "Assets = Liabilities + Equity?", indent=2, bold=True)
    for yr in years:
        c = yr_col[yr]
        check_cell = ws.cell(row, c)
        check_cell.value = (
            f'=IF(ABS({_col(c)}{total_assets_row}-{_col(c)}{total_le_row})<1,"✓ BALANCED",'
            f'"ERROR: "&TEXT(ABS({_col(c)}{total_assets_row}-{_col(c)}{total_le_row}),"$#,##0"))'
        )
        check_cell.font = check_cell.font.copy() if check_cell.font else None

        ok_fill = PatternFill(fill_type=FILL_SOLID, fgColor=cfg.FILL_CHECK_OK)
        fail_fill = PatternFill(fill_type=FILL_SOLID, fgColor=cfg.FILL_CHECK_FAIL)
        cell_ref = f"{_col(c)}{row}"
        ws.conditional_formatting.add(
            cell_ref,
            FormulaRule(formula=[f'LEFT({cell_ref},1)="✓"'], fill=ok_fill)
        )
        ws.conditional_formatting.add(
            cell_ref,
            FormulaRule(formula=[f'LEFT({cell_ref},1)="E"'], fill=fail_fill)
        )
    ws.row_dimensions[row].height = cfg.ROW_HEIGHT_NORMAL
    check_row = row
    row += 1

    # ── Column widths & display ────────────────────────────────────────────────
    set_standard_column_widths(ws, n)
    freeze_header(ws, freeze_row=4)
    set_print_landscape(ws)

    # ── Named ranges ──────────────────────────────────────────────────────────
    _register_named_ranges(wb, ws, years, yr_col,
                           cash_row=cash_row, total_assets_row=total_assets_row,
                           total_liab_row=total_liab_row, total_equity_row=total_equity_row,
                           tca_row=tca_row, tcl_row=tcl_row, ar_row=ar_row,
                           inv_row=inv_row, ap_row=ap_row, ltd_row=ltd_row,
                           std_row=std_row, sti_row=sti_row)


def _register_named_ranges(wb: Workbook, ws, years, yr_col, **row_map) -> None:
    from openpyxl.workbook.defined_name import DefinedName
    sn = f"'{cfg.SHEET_BS}'"
    for i, yr in enumerate(years):
        c = get_column_letter(yr_col[yr])
        suffix = f"_FY{i + 1}"
        for name, row_key in [
            ("BS_Cash", "cash_row"),
            ("BS_TotalAssets", "total_assets_row"),
            ("BS_TotalLiab", "total_liab_row"),
            ("BS_TotalEquity", "total_equity_row"),
            ("BS_TCA", "tca_row"),
            ("BS_TCL", "tcl_row"),
            ("BS_AR", "ar_row"),
            ("BS_Inventory", "inv_row"),
            ("BS_AP", "ap_row"),
            ("BS_LTD", "ltd_row"),
            ("BS_STD", "std_row"),
            ("BS_STI", "sti_row"),
        ]:
            r = row_map[row_key]
            ref = f"{sn}!${c}${r}"
            try:
                wb.defined_names[name + suffix] = DefinedName(name + suffix, attr_text=ref)
            except Exception:
                pass
