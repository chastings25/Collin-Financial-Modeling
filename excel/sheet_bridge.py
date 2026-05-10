"""Build the EBITDA Bridge worksheet (waterfall table for most recent fiscal year)."""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.formatting.rule import DataBarRule

import config as cfg
from models.financial_data import FinancialData
from excel.styles import (
    style_title, style_col_header, style_subsection_header,
    style_label, style_subtotal, style_grand_total, style_pct,
    right_align, set_print_landscape,
)

_COL_LABEL = 1
_COL_AMOUNT = 2
_COL_PCT_REV = 3
_COL_BAR = 4


def build_ebitda_bridge_sheet(wb: Workbook, data: FinancialData) -> None:
    ws = wb[cfg.SHEET_BRIDGE]
    ws.sheet_properties.tabColor = cfg.TAB_BRIDGE

    years = data.sorted_years()
    if not years:
        ws.cell(1, 1).value = "No data available"
        return

    fy = len(years)  # most recent FY index (1-based named range suffix)
    yr = years[-1]
    meta = data.metadata

    # ── Header ─────────────────────────────────────────────────────────────────
    style_title(ws.cell(1, 1), meta.company_name)
    for c in range(2, 5):
        style_title(ws.cell(1, c))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws.row_dimensions[1].height = cfg.ROW_HEIGHT_TITLE

    style_subsection_header(ws.cell(2, 1), f"EBITDA Bridge — FY{yr}  ($ in {meta.units})")
    for c in range(2, 5):
        style_subsection_header(ws.cell(2, c))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    ws.row_dimensions[2].height = cfg.ROW_HEIGHT_HEADER

    style_col_header(ws.cell(3, 1), "Component")
    style_col_header(ws.cell(3, 2), f"Amount ($ {meta.units})")
    style_col_header(ws.cell(3, 3), "% of Revenue")
    style_col_header(ws.cell(3, 4), "Visual")
    ws.row_dimensions[3].height = cfg.ROW_HEIGHT_HEADER

    row = 4
    bar_rows: list[int] = []

    def write_row(label: str, amount_formula: str, pct_formula: str | None,
                  row_style: str = "link", indent: int = 2) -> int:
        nonlocal row
        style_label(ws.cell(row, _COL_LABEL), label, indent=indent,
                    bold=(row_style in ("subtotal", "total")))

        amt = ws.cell(row, _COL_AMOUNT)
        amt.value = amount_formula
        amt.number_format = cfg.FMT_CURRENCY
        amt.alignment = right_align()
        if row_style == "total":
            style_grand_total(amt)
        elif row_style == "subtotal":
            style_subtotal(amt)
        else:
            amt.font = Font(color=cfg.COLOR_LINK, name=cfg.FONT_NAME, size=cfg.FONT_SIZE_NORMAL)

        if pct_formula:
            pct = ws.cell(row, _COL_PCT_REV)
            pct.value = pct_formula
            pct.number_format = cfg.FMT_PCT
            pct.alignment = right_align()

        bar = ws.cell(row, _COL_BAR)
        bar.value = f"=IF(ISNUMBER(B{row}),ABS(B{row}),0)"
        bar.number_format = cfg.FMT_CURRENCY
        bar_rows.append(row)

        ws.row_dimensions[row].height = cfg.ROW_HEIGHT_NORMAL
        r = row
        row += 1
        return r

    def spacer() -> None:
        nonlocal row
        ws.row_dimensions[row].height = cfg.ROW_HEIGHT_SPACER
        row += 1

    def safe_pct(amount_expr: str) -> str:
        rev = f"IS_Revenue_FY{fy}"
        return (
            f"=IF(AND(ISNUMBER({rev}),{rev}<>0),"
            f"({amount_expr})/{rev},\"\")"
        )

    # ── Bridge rows ───────────────────────────────────────────────────────────
    write_row("Revenue",
              f"=IS_Revenue_FY{fy}",
              safe_pct(f"IS_Revenue_FY{fy}"),
              "subtotal")

    write_row("(−) Cost of Revenue",
              f"=-ABS(IS_COGS_FY{fy})",
              safe_pct(f"-ABS(IS_COGS_FY{fy})"))

    write_row("= Gross Profit",
              f"=IS_GrossProfit_FY{fy}",
              safe_pct(f"IS_GrossProfit_FY{fy}"),
              "subtotal")
    spacer()

    # Operating Expenses (excl. D&A) derived from EBITDA bridge identity:
    # EBITDA = Gross Profit - OpEx + D&A  →  OpEx = Gross Profit - EBITDA + D&A  (when positive)
    # We show the negative impact: -(GrossProfit - EBITDA - DA)
    write_row(
        "(−) Operating Expenses (excl. D&A)",
        (f"=IF(AND(ISNUMBER(IS_GrossProfit_FY{fy}),ISNUMBER(IS_EBITDA_FY{fy}),"
         f"ISNUMBER(IS_DA_FY{fy})),"
         f"-(IS_GrossProfit_FY{fy}-IS_EBITDA_FY{fy}-IS_DA_FY{fy}),\"\")"),
        (f"=IF(AND(ISNUMBER(IS_GrossProfit_FY{fy}),ISNUMBER(IS_EBITDA_FY{fy}),"
         f"ISNUMBER(IS_DA_FY{fy}),ISNUMBER(IS_Revenue_FY{fy}),IS_Revenue_FY{fy}<>0),"
         f"-(IS_GrossProfit_FY{fy}-IS_EBITDA_FY{fy}-IS_DA_FY{fy})/IS_Revenue_FY{fy},\"\")"),
    )

    write_row(
        "(+) Depreciation & Amortization Add-back",
        f"=IF(ISNUMBER(IS_DA_FY{fy}),IS_DA_FY{fy},\"\")",
        f"=IF(AND(ISNUMBER(IS_DA_FY{fy}),ISNUMBER(IS_Revenue_FY{fy}),IS_Revenue_FY{fy}<>0),"
        f"IS_DA_FY{fy}/IS_Revenue_FY{fy},\"\")",
    )
    spacer()

    write_row("= EBITDA",
              f"=IS_EBITDA_FY{fy}",
              safe_pct(f"IS_EBITDA_FY{fy}"),
              "total")
    spacer()

    # EBIT continuation
    write_row("(−) Depreciation & Amortization",
              f"=IF(ISNUMBER(IS_DA_FY{fy}),-IS_DA_FY{fy},\"\")",
              f"=IF(AND(ISNUMBER(IS_DA_FY{fy}),ISNUMBER(IS_Revenue_FY{fy}),IS_Revenue_FY{fy}<>0),"
              f"-IS_DA_FY{fy}/IS_Revenue_FY{fy},\"\")")

    write_row("= EBIT (Operating Income)",
              f"=IS_EBIT_FY{fy}",
              safe_pct(f"IS_EBIT_FY{fy}"),
              "subtotal")

    # ── DataBar on bar column ──────────────────────────────────────────────────
    if bar_rows:
        bar_range = f"D{bar_rows[0]}:D{bar_rows[-1]}"
        ws.conditional_formatting.add(
            bar_range,
            DataBarRule(
                start_type="min", start_value=0,
                end_type="max", end_value=None,
                color="FF4472C4",
            ),
        )

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = cfg.COL_WIDTH_LABEL
    ws.column_dimensions["B"].width = cfg.COL_WIDTH_YEAR
    ws.column_dimensions["C"].width = cfg.COL_WIDTH_YEAR
    ws.column_dimensions["D"].width = 24

    set_print_landscape(ws)
